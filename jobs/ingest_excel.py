"""Parse the curation Excel (data/Curadoria Engelstein.v65.xlsx) into RestaurantDocs.

Parsing only — writing to Cosmos belongs to jobs.seed_restaurants, which merges
these docs with the Google Maps resolution first.

Usage:
    uv run python -m jobs.ingest_excel stats   # parse + report
    uv run python -m jobs.ingest_excel notes    # dedupe decisions

Same-name rows are merged when compatible (complementary halves of one place);
rows with conflicting scalar data are kept as distinct restaurants with a
numeric slug suffix — canonical identity arrives with the Maps resolve job.
"""

import json
import re
from collections import Counter
from pathlib import Path

import typer
from openpyxl import load_workbook

from core.cuisines import canonical_cuisines
from core.models.restaurant import AwardMention, RestaurantDoc
from jobs.fixtures import make_restaurant_id

app = typer.Typer(help="Excel → canonical restaurant docs")

DEFAULT_XLSX = Path(__file__).resolve().parents[2] / "data" / "Curadoria Engelstein.v65.xlsx"
BASE_SHEET = "Base Couvert"
XLSX_OPTION = typer.Option(DEFAULT_XLSX, exists=True)
DRY_RUN_OPTION = typer.Option(False, "--dry-run")

# Cuisine spelling variants live in core.cuisines — the API resolves query labels
# through the same table, so ingestion and filtering can't drift apart.

_AWARD_ENTRY_RE = re.compile(r"^(?P<placement>[^|]+)\|\s*(?P<rest>.+)$")
_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def split_multi(value: object) -> list[str]:
    """Semicolon-separated cell → deduped list, order preserved."""
    if value is None:
        return []
    seen: dict[str, None] = {}
    for part in str(value).split(";"):
        item = part.strip()
        if item:
            seen.setdefault(item, None)
    return list(seen)


def normalize_cuisines(values: list[str]) -> list[str]:
    return canonical_cuisines(values)


def parse_awards_mention(value: object) -> list[AwardMention]:
    mentions = []
    for entry in split_multi(value):
        m = _AWARD_ENTRY_RE.match(entry)
        if m:
            rest = m.group("rest").strip()
            category, sep, award = rest.rpartition(" - ")
            if not sep:
                category, award = "", rest
            year_match = _YEAR_RE.search(award)
            mentions.append(
                AwardMention(
                    raw=entry,
                    placement=m.group("placement").strip(),
                    category=category.strip() or None,
                    award=award.strip() or None,
                    year=int(year_match.group()) if year_match else None,
                )
            )
        else:
            mentions.append(AwardMention(raw=entry))
    return mentions


def _clean(value: object) -> str | None:
    text = str(value).strip() if value is not None else ""
    return text or None


def row_to_doc(row: tuple) -> RestaurantDoc | None:
    """Columns: (blank, Nº, Nome, Tipo, Ocasião, Cozinha, Posicionamento, Chef,
    Menção em Premiações, Especialidades, Comentários Editoriais)."""
    name = _clean(row[2])
    if name is None:
        return None
    source_row = None
    if row[1] is not None and str(row[1]).strip().isdigit():
        source_row = int(str(row[1]).strip())
    return RestaurantDoc(
        id=make_restaurant_id(name),
        name=name,
        source_name=name,
        source_row=source_row,
        place_types=split_multi(row[3]),
        occasions=split_multi(row[4]),
        cuisines=normalize_cuisines(split_multi(row[5])),
        positionings=split_multi(row[6]),
        chef=_clean(row[7]),
        awards_mention=_clean(row[8]),
        awards=parse_awards_mention(row[8]),
        specialties=split_multi(row[9]),
        editorial_comment=_clean(row[10]),
    )


_MERGE_SCALARS = ("chef", "awards_mention", "editorial_comment")
_MERGE_LISTS = ("place_types", "occasions", "cuisines", "positionings", "specialties")


def _compatible(a: RestaurantDoc, b: RestaurantDoc) -> bool:
    """Same place if no scalar field disagrees where both rows have data."""
    for field in _MERGE_SCALARS:
        va, vb = getattr(a, field), getattr(b, field)
        if va is not None and vb is not None and va != vb:
            return False
    return True


def _merge(a: RestaurantDoc, b: RestaurantDoc) -> RestaurantDoc:
    update: dict = {}
    for field in _MERGE_SCALARS:
        if getattr(a, field) is None and getattr(b, field) is not None:
            update[field] = getattr(b, field)
    for field in _MERGE_LISTS:
        seen: dict[str, None] = {}
        for item in getattr(a, field) + getattr(b, field):
            seen.setdefault(item, None)
        update[field] = list(seen)
    if not a.awards and b.awards:
        update["awards"] = b.awards
    return a.model_copy(update=update)


def dedupe(docs: list[RestaurantDoc]) -> tuple[list[RestaurantDoc], list[str]]:
    by_id: dict[str, RestaurantDoc] = {}
    notes: list[str] = []
    for doc in docs:
        slug = doc.id
        if slug not in by_id:
            by_id[slug] = doc
            continue
        if _compatible(by_id[slug], doc):
            by_id[slug] = _merge(by_id[slug], doc)
            notes.append(f"merged duplicate rows for '{doc.name}' ({slug})")
        else:
            n = 2
            while f"{slug}-{n}" in by_id:
                n += 1
            distinct = doc.model_copy(update={"id": f"{slug}-{n}"})
            by_id[distinct.id] = distinct
            notes.append(
                f"CONFLICT: '{doc.name}' rows disagree — kept as {slug} and {distinct.id}"
            )
    return list(by_id.values()), notes


def parse_workbook(path: Path = DEFAULT_XLSX) -> tuple[list[RestaurantDoc], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = wb[BASE_SHEET].iter_rows(min_row=3, values_only=True)
    docs = [doc for row in rows if (doc := row_to_doc(row)) is not None]
    return dedupe(docs)


@app.command()
def stats(xlsx: Path = XLSX_OPTION) -> None:
    """Parse and report — makes no network calls and writes nothing."""
    docs, notes = parse_workbook(xlsx)
    typer.echo(f"restaurants: {len(docs)}")
    typer.echo(f"with awards: {sum(1 for d in docs if d.awards)}")
    typer.echo(f"with chef: {sum(1 for d in docs if d.chef)}")
    cuisines = Counter(c for d in docs for c in d.cuisines)
    typer.echo(f"cuisines ({len(cuisines)}): {cuisines.most_common(12)}")
    for note in notes:
        typer.echo(f"  {note}")
    sample = docs[0].model_dump()
    typer.echo(json.dumps(sample, ensure_ascii=False, indent=1))


@app.command()
def notes(xlsx: Path = XLSX_OPTION) -> None:
    """Merge/conflict decisions taken while deduping same-name rows."""
    _, messages = parse_workbook(xlsx)
    for message in messages:
        typer.echo(f"  {message}")
    typer.echo(f"{len(messages)} note(s)")


if __name__ == "__main__":
    app()
